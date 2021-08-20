# Author: Quinn James (qj@quinnjam.es)
# 
# A program for updating an old-format excel workbook into a current-format
# workbook.
# 
# Copies data from a .xls fle into a .xlsx file. Changes numerical times and
# dates into excel-format times and dates
#
# More details about this project can be found in the README file or at:
#   https://github.com/qcjames53/AJM-RouteSummaries

import win32com.client
import openpyxl
import math
import datetime

from Log import Log

# CONSTANTS
DEBUG = True    # Whether Excel should be visible when updating document format
VALID_HEADERS = [
    ["sequence", "seq"],
    ["date"],
    ["route"],
    ["direction", "dir"],
    ["run"],
    ["start time", "start"],
    ["onboard", "ob"],
    ["stop number", "stop no", "stp number", "stp no", "stop", "stp"],
    ["arrival time", "arrive time", "arrival", "arrive"],
    ["schedule time", "sched time", "schedule", "sched"],
    ["offs", "departures", "departure"],
    ["ons", "arrivals", "arrival"],
    ["loads", "load", "ld"],
    ["time check", "time chk", "time", "check"]
]

class ExcelApplication:
    def __init__(self):
        self.app = win32com.client.Dispatch('Excel.Application')
        self.app.visible = DEBUG

    def __del__(self):
        '''
        Closes the Excel application whenever this object is deleted. Used for
        convenience and in case of program crash.
        '''
        self.app.Quit()

    def convertToXLSX(self, old_filepath, new_filepath):
        # Open the old book
        wb = self.app.WorkBooks.Open(old_filepath)

        # Save the new book
        self.app.DisplayAlerts = False
        wb.SaveAs(new_filepath, FileFormat = 51, ConflictResolution=2)
        wb.Close()
        self.app.DisplayAlerts = True

def convertFormat(log:Log, old_filepath, new_filepath) -> None:
    '''
    Updates an excel document from .xls to .xlsx format without modifying
    the document contents.

    @param old_filepath The filepath for the document to open
    @param new_filepath The filepath for the document to save
    '''
    # Update the document format from .xls to .xlsx
    excelApp = ExcelApplication()
    excelApp.convertToXLSX(old_filepath, new_filepath)

def convertValues(log:Log, filepath, save_filepath=None) -> None:
    '''
    Updates the values of a .xlsx workbook to match the required format.
    Numerical dates and times become excel-format dates and times.

    @param filepath The filepath of the excel document.
    @param save_filepath Optional: The filepath to save the document to
    '''
    if save_filepath is None:
        save_filepath = filepath

    # Load the file
    try:
        wb = openpyxl.load_workbook(filename=filepath)
    except:
        log.logError("Could not load the workbook")
        return
    ride_checks = wb.active

    # Check that the headers are correct, warn if not
    for col in range(1, 15):
        cell_string = str(ride_checks.cell(row=1, column=col).value)
        if cell_string.lower() not in VALID_HEADERS[col-1]:
            col_letter = openpyxl.utils.get_column_letter(col)
            log.logWarning(f"The value for column {col_letter}, '{cell_string}', is not a valid header value. Please check that this column contains data representing the {VALID_HEADERS[col-1][0]}")

    # Go through sheet until out of rows, skipping header
    row = 2
    while ride_checks.cell(row=row, column=1).value is not None:
        # Change date column
        date_no = ride_checks.cell(row=row, column=2).value
        if isinstance(date_no, int):
            year = date_no % 100
            if year < 100:
                year = year + 2000
            month = math.floor(date_no / 10000) % 100
            day = math.floor(date_no / 100) % 100
            if 1900 > year or year > 2100 or 1 > month or month > 12 or \
                1 > day or day > 31:
                log.logWarning(f"Row {row}: bad date calculated - {year:04}-{month:02}-{day:02}. Check input data.")
            else:
                date = datetime.date(year, month, day)
                ride_checks.cell(row=row, column=2).value = date
        else:
            log.logWarning(f"Row {row}: The date is not an integer. Check input data")

        # Change start time column
        time_no = ride_checks.cell(row=row, column=6).value
        if isinstance(time_no, int):
            hour = math.floor(time_no / 100) % 100
            if hour > 23:
                hour %= 24
            minute = time_no % 100
            if 0 > hour or 23 < hour or 0 > minute or 59 < minute:
                log.logWarning(f"Row {row}: bad times calculated - {hour:02}:{minute:02}. Check input data.")
            else:
                time = datetime.time(hour, minute)
                ride_checks.cell(row=row, column=6).value = time
        else:
            log.logWarning(f"Row {row}: The start time is not an integer. Check input data")

        # Change arrival time column, if cell is not empty
        time_no = ride_checks.cell(row=row, column=9).value
        if (time_no is not None) and (time_no != ""):
            if isinstance(time_no, int):
                hour = math.floor(time_no / 100) % 100
                if hour > 23:
                    hour %= 24
                minute = time_no % 100
                if 0 > hour or 23 < hour or 0 > minute or 59 < minute:
                    log.logWarning(f"Row {row}: bad times calculated - {hour:02}:{minute:02}. Check input data.")
                else:
                    time = datetime.time(hour, minute)
                    ride_checks.cell(row=row, column=9).value = time
            else:
                log.logWarning(f"Row {row}: The arrival time is not an integer. Check input data")

        # Change schedule time column, if cell is not empty
        time_no = ride_checks.cell(row=row, column=10).value
        if (time_no is not None) and (time_no != ""):
            if isinstance(time_no, int):
                hour = math.floor(time_no / 100) % 100
                if hour > 23:
                    hour %= 24
                minute = time_no % 100
                if 0 > hour or 23 < hour or 0 > minute or 59 < minute:
                    log.logWarning(f"Row {row}: bad times calculated - {hour:02}:{minute:02}. Check input data.")
                else:
                    time = datetime.time(hour, minute)
                    ride_checks.cell(row=row, column=10).value = time
            else:
                log.logWarning(f"Row {row}: The schedule time is not an integer. Check input data")

        # Increment
        row += 1
    
    # Try to save the output file
    try:
        wb.save(save_filepath)
        log.logGeneral("Succesfully updated the workbook.")
    except:
        log.logFailure("Could not save the workbook.")