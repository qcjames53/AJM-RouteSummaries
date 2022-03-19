# Author: Quinn James (qj@quinnjam.es)
#
# A command-line utility to generate a ridership summary for bus routes.
# Functionality of this program's input and output files are documented in the
# main README.md file.
#
# More details about this project can be found in the README file or at:
#   https://github.com/qcjames53/AJM-RouteSummaries

import openpyxl
from openpyxl.styles import Alignment
import datetime
from enum import Enum

from Log import Log

# Route direction enum
class Direction(Enum):
    IB = "IB" # inbound
    OB = "OB" # outbound
    NB = "NB" # northbound
    SB = "SB" # southbound
    EB = "EB" # eastbound
    WB = "WB" # westbound
    LP = "LP" # loop
    UN = "UN" # unknown

# Utility functions
def stringToDirection(dir_string) -> Direction:
    """
    Converts a direction string to a direction enum

    @param dir_string The string to convert
    @returns Direction enum for the direction of this string
    """
    direction = Direction.UN
    if dir_string == "IB":
        direction = Direction.IB
    elif dir_string == "OB":
        direction = Direction.OB
    elif dir_string == "NB":
        direction = Direction.NB
    elif dir_string == "SB":
        direction = Direction.SB
    elif dir_string == "EB":
        direction = Direction.EB
    elif dir_string == "WB":
        direction = Direction.WB
    elif dir_string == "LP":
        direction = Direction.LP
    return direction


class RouteManager:
    """ 
    A class that allows for easy storage of routes and their respective data
    """
    def __init__(self, log:Log) -> None:
        """
        Initilize RouteManager with the appropriate internal variables
        
        @param log The log object in the currect workbook
        """
        # A map of routes from routestring to route object
        self.routes = {}
        self.log = log

    def __str__(self) -> str:
        output = ""
        for key in sorted(self.routes.keys()):
            output += self.routes[key].__str__() + '\n'
        return output

    def __repr__(self) -> str:
        return self.__str__()

    def addStop(self, route, stop_no, street, cross_street) -> None:
        """
        Adds a stop to the specified route object.
        
        @param route The route with which to add the stop
        @param stop_no The stop number to add
        @param street The main street for the stop
        @param cross_street The cross street for the stop
        """
        # if the route key does not exist, create the route
        if route not in self.routes:
            self.routes[route] = Route(route, self.log)
        
        # Add the stop to the appropriate route
        self.routes[route].addStop(stop_no, street, cross_street)

    def addData(self, route, stop_no, datetime, run, arrival_time, \
        schedule_time, offs, ons, onboard) -> bool:
        """
        Adds data about a stop to the specified route object

        @param route The route with which to add the stop
        @param stop_no The stop number
        @param datetime The datetime the route began
        @param run The run value of this stop
        @param arrival_time The arrival time of the stop
        @param schedule_time The scheduled arrival time of the stop
        @param offs The number of passengers departing the bus
        @param ons The number of passengers boarding the bus
        @param onboard The number of passengers carrying over from a prev route
        @returns Boolean of whether the data was successfully added
        """
        # If the route does not exist, log an error and return
        if route not in self.routes:
            self.log.logError("Tried to add data to nonexistent route: " + str(route))
            return False

        # Add the data to the appropriate route
        return self.routes[route].addData(stop_no, datetime, run, arrival_time,\
            schedule_time, offs, ons, onboard)

    def setRouteData(self, route, description, direction: Direction) -> None:
        """
        Sets the metadata of a particular route.
        
        @param route The number of the route
        @param description A text description of the route (University, Uptown)
        @param direction The direction of the route as a Direction object
        """
        # if the route key does not exist, create the route
        if route not in self.routes:
            self.routes[route] = Route(route, self.log)

        # Add data to appropriate route
        self.routes[route].setRouteData(description, direction)

    def buildLoad(self) -> None:
        """
        Builds and saves the loads for all routes
        """
        for route in self.routes:
            self.routes[route].buildLoad()

    def buildRouteTotals(self, worksheet) -> None:
        """
        Builds a worksheet of route totals. See the README file for further description regarding this functionality.

        @param worksheet The excel worksheet to operate on
        """
        # Display headers
        worksheet["A1"] = "Route #"
        worksheet["A1"].alignment = Alignment(horizontal="right")
        worksheet["C1"] = "Route"
        worksheet["D1"] = "Ons"
        worksheet["D1"].alignment = Alignment(horizontal="right")
        worksheet["E1"] = "Offs"
        worksheet["E1"].alignment = Alignment(horizontal="right")
        worksheet["F1"] = "Total"
        worksheet["F1"].alignment = Alignment(horizontal="right")

        # Set the width of the columns
        worksheet.column_dimensions["A"].width = 8
        worksheet.column_dimensions["B"].width = 1
        worksheet.column_dimensions["C"].width = 30
        
        # Display values
        current_row = 2
        for route in sorted(self.routes.keys()):
            offs, ons, total = self.routes[route].getTotalOffsAndOns()
            worksheet.cell(row=current_row, column=1).value = route
            worksheet.cell(row=current_row, column=3).value = \
                self.routes[route].getDescriptorAndDirectionTrunc(29)
            worksheet.cell(row=current_row, column=4).value = ons
            worksheet.cell(row=current_row, column=5).value = offs
            worksheet.cell(row=current_row, column=6).value = total

            current_row += 1


    def buildMaxLoads(self, worksheet) -> None:
        """
        Builds a worksheet of max loads for each route. See the README file for 
        further description regarding this functionality.

        @param worksheet The excel worksheet to operate on
        """
        # Display headers
        worksheet["A1"] = "Route No."
        worksheet["B1"] = "Route"
        worksheet["C1"] = "Start Time"
        worksheet["D1"] = "Ons"
        worksheet["E1"] = "Offs"
        worksheet["F1"] = "Max Load"

        # Display values
        current_row = 2
        for route in sorted(self.routes.keys()):
            current_row = self.routes[route].buildTotalsByTime(worksheet,\
            current_row)

    def buildRouteTotalsByStop(self, worksheet) -> None:
        """
        Builds a worksheet of route totals for each stop. See the README file
        for further description regarding this functionality.

        @param worksheet The excel worksheet to operate on
        """
        worksheet["A1"] = "Route"
        worksheet["B1"] = "Route Name"
        worksheet["C1"] = "Stop"
        worksheet["D1"] = "Street"
        worksheet["E1"] = "Cross Street"
        worksheet["F1"] = "Ons"
        worksheet["G1"] = "Offs"
        worksheet["H1"] = "Total"
        worksheet["I1"] = "Load"

        # Display values
        current_row = 2
        for route in sorted(self.routes.keys()):
            current_row = self.routes[route].buildRouteTotalsByStop(worksheet, \
                current_row)

    def buildOnTimeDetail(self, worksheet) -> None:
        """
        Builds a worksheet of whether busses were on time or not. See the README
        file for further description regarding this functionality.

        @param worksheet The excel worksheet to operate on
        """
        worksheet["A1"] = "Route"
        worksheet["B1"] = "Route Name"
        worksheet["C1"] = "Date"
        worksheet["D1"] = "Time"
        worksheet["E1"] = "Run"

        worksheet.column_dimensions["C"].width = 11

        # Display values
        current_row = 2
        for route in sorted(self.routes.keys()):
            current_row = self.routes[route].buildOnTimeDetail(worksheet, \
                current_row)

    def buildDetailReport(self, worksheet) -> None:
        """
        Builds a detailed report of all data collected. See the README file for 
        further details regarding this functionality.

        @param worksheet The excel worksheet to operate on
        """
        current_row = 1
        for route in sorted(self.routes.keys()):
            current_row = self.routes[route].buildDetailReport(worksheet, \
                current_row)


class Route:
    """
    A class that allows for easy storage of the stops inside of a route
    """
    def __init__(self, route, log:Log) -> None:
        """
        Initialize Route with the appropriate internal variables
        
        @param route The route number
        @param log The workbook's log object
        """
        self.route = route
        self.stops = {}
        self.times = []
        self.descriptor = "Descriptor Unset"
        self.direction = Direction.UN
        self.log = log
        self.timed_stops = []
        self.onboard = {}

    def __str__(self) -> str:
        output = ""
        for key in sorted(self.stops.keys()):
            output += str(self.stops[key])
        return output

    def __repr__(self) -> str:
        return self.__str__()

    def addStop(self, stop_no, street, cross_street) -> None:
        """
        Adds a stop of stop_no to this route
        
        @param stop_no The number of the stop.
        @param street The main street of this stop.
        @param cross_street The main cross street of this stop.
        """
        # If stop exists, throw error and return
        if stop_no in self.stops:
            self.log.logError("Tried to add stop " + str(stop_no) + " to route " + str(self.route) + " when it already exists.")
            return

        self.stops[stop_no] = Stop(self.route, stop_no, street, cross_street)
        self.stops[stop_no].setRouteData(self.descriptor, self.direction)

    def addData(self, stop_no, datetime, run, arrival_time, schedule_time, \
        offs, ons, onboard) -> bool:
        """
        Adds data about a stop to a specific stop object
        
        @param stop_no The stop number where we're adding this data
        @param datetime The datetime when the route began
        @param run The run value for this stop
        @param arrival_time Optional: The arrival time for this stop
        @param schedule_time Optional: The scheduled arrival time for this stop
        @param offs Optional: The number of passengers departing the bus
        @param ons Optional: The number of passengers boarding the bus
        @param onboard The number of passengers carrying over from a prev route
        @returns Boolean of whether the data was successfully added
        """
        # If stop_no does not exist, throw error and return
        if stop_no not in self.stops:
            self.log.logError("Tried to add data to stop " + str(stop_no) + " in route " + str(self.route) + " when stop does not exist.")
            return False

        # If datetime does not exist, add it to datetimes and sort
        if datetime not in self.times:
            self.times.append(datetime)

            # Sort by time, date instead of date, time
            self.times.sort(key=lambda x: (x.time(), x.date()))

        # If has time data, add to timed stops where neccesary
        if (arrival_time is not None) and (stop_no not in self.timed_stops):
            self.timed_stops.append(stop_no)
            self.timed_stops.sort()

        # Set the onboard value for this route if the value is provided
        if onboard is not None:
            # Log a warning if we're changing an already set onboard number
            if datetime in self.onboard and self.onboard[datetime] != onboard:
                self.log.logWarning("Route " + str(self.route) + " time " + \
                    str(datetime) + " stop " + str(stop_no) + \
                    "Overriding onboard value " + str(self.onboard[datetime]) +\
                    " with new value " + str(onboard))
            # Set the value regardless
            self.onboard[datetime] = onboard

        return self.stops[stop_no].addData(datetime, run, arrival_time, \
            schedule_time, offs, ons)

    def setRouteData(self, description, direction: Direction) -> None:
        """
        Sets the metadata for this route
        
        @param description A text description of the route (University, Uptown)
        @param direction The direction of the route as a Direction object
        """
        self.descriptor = description
        self.direction = direction

        # Set for all child stops
        for stop_no in self.stops:
            self.stops[stop_no].setRouteData(description, direction)

    def buildLoad(self) -> None:
        """
        Builds and saves the loads for all data points within this route
        """
        times_by_datetime = sorted(self.times)
        current_load = 0
        for datetime in times_by_datetime:
            # Set the starting load to the onboard value (stored)
            current_load = 0
            if datetime in self.onboard:
                current_load = self.onboard[datetime]

            for stop_no in sorted(self.stops.keys()):
                current_off, current_on = \
                    self.stops[stop_no].getOffsAndOns(datetime)
                current_load = current_load + current_on - current_off
                self.stops[stop_no].setLoad(datetime, current_load)

                # Display an error if current_load drops below 0
                if current_load < 0:
                    self.log.logWarning("Route " + str(self.route) + " time " +\
                        str(datetime) + " stop " + str(stop_no) + \
                        ": The load has dropped below 0 (check for bad data)")

    def getDescriptorAndDirection(self) -> str:
        """
        @returns A string representation of this routes descriptor and direction
        """
        return self.descriptor + " " + str(self.direction.value)

    def getDescriptorAndDirectionTrunc(self, l) -> str:
        """
        @param l Length of the output string

        @returns A truncated string representation of this routes descriptor and direction
        """
        remaining_l = max(0, l - len(str(self.direction.value)) - 1)
        return self.descriptor[0:remaining_l] + " " + str(self.direction.value)

    def getTotalOffsAndOns(self) -> tuple:
        """
        @returns the total offs, ons, and total passengers for this route over 
            all stops
        """
        offs = 0
        ons = 0
        for stop_no in self.stops:
            current_off, current_on = self.stops[stop_no].getTotalOffsAndOns()
            offs += current_off
            ons += current_on
        total = offs + ons
        return offs, ons, total

    def buildTotalsByTime(self, worksheet, current_row) -> int:
        """
        Output totals for every time of every route into a provided workbook
        
        @param worksheet The workbook to output to
        @param current_row The row of the worksheet to start on
        @returns The next empty row in the worksheet
        """
        # Loop over every datetime
        dt_search_index = 0
        while dt_search_index < len(self.times):
            # Get a list of all datetimes with the same start time as the
            # current index
            search_datetimes = [self.times[dt_search_index]]
            dt_search_index += 1
            while dt_search_index < len(self.times) and \
                search_datetimes[0].time() == \
                self.times[dt_search_index].time():
                search_datetimes.append(self.times[dt_search_index])
                dt_search_index += 1

            # Generate totals and max for all stops for all datetimes in list
            offs = 0
            ons = 0
            current_load = 0
            max_load = 0
            for datetime in search_datetimes:
                for stop_no in self.stops:
                    current_off, current_on, current_load = \
                        self.stops[stop_no].getOffsOnsAndLoad(datetime)

                    # Add to tallies
                    offs += current_off
                    ons += current_on

                    # Save max load if is the current max
                    if current_load > max_load:
                        max_load = current_load

            # Write data to sheet
            worksheet.cell(row=current_row, column=1).value = self.route
            worksheet.cell(row=current_row, column=2).value = \
                self.getDescriptorAndDirection()
            worksheet.cell(row=current_row, column=3).value = \
                datetime.strftime("%H:%M")
            worksheet.cell(row=current_row, column=4).value = ons
            worksheet.cell(row=current_row, column=5).value = offs
            worksheet.cell(row=current_row, column=6).value = max_load

            current_row += 1
        return current_row

    def buildRouteTotalsByStop(self, worksheet, current_row) -> int:
        """
        Builds total values for the route for each stop.

        @param worksheet The excel worksheet to operate on
        @param current_row The row of the worksheet to start on
        @returns The next empty row in the worksheet
        """
        for stop_no in self.stops:
            current_row = self.stops[stop_no].buildRouteTotalsByStop( \
                worksheet, current_row)
        return current_row

    def buildOnTimeDetail(self, worksheet, current_row) -> int:
        """
        Builds table of stops that are on time or not.

        @param worksheet The excel worksheet to operate on
        @param current_row The row of the worksheet to start on
        @returns The next empty row in the worksheet
        """
        # If there are no timed stops, skip this stop
        if len(self.timed_stops) == 0:
            return current_row

        # Build header
        col = 6
        for stop in self.timed_stops:
            worksheet.cell(row=current_row, column=col).value = \
                self.stops[stop].street
            worksheet.cell(row=current_row + 1, column=col).value = \
                self.stops[stop].cross_street
            col += 1
        current_row += 2

        # Generate one row per time
        for datetime in self.times:
            # Write route data to row
            worksheet.cell(row=current_row, column=1).value = self.route
            worksheet.cell(row=current_row, column=2).value = \
                self.getDescriptorAndDirection()
            worksheet.cell(row=current_row, column=3).value = datetime.date()
            worksheet.cell(row=current_row, column=4).value = datetime.time()
            
            # Write run number to row
            worksheet.cell(row=current_row, column=5).value = \
                self.stops[self.timed_stops[0]].getRun(datetime)

            # Write stop data to row
            col = 6
            for stop in self.timed_stops:
                minutes_late = self.stops[stop].getMinutesLate(datetime)
                # Handle error cases
                if minutes_late is None:
                    worksheet.cell(row=current_row, column=col).value = "NA"
                else:
                    worksheet.cell(row=current_row, column=col).value = \
                        minutes_late
                col += 1

            current_row += 1
        return current_row + 1

    def buildDetailReport(self, worksheet, current_row) -> int:
        """
        Builds a detailed report of all data collected.

        @param worksheet The excel worksheet to operate on
        @param current_row The worksheet row to start on
        @returns The next empty row in the worksheet
        """
        # Display the header
        worksheet.cell(row=current_row, column=3).value = "Route #"
        worksheet.cell(row=current_row, column=4).value = self.route
        worksheet.cell(row=current_row, column=5).value = \
            self.getDescriptorAndDirection()
        worksheet.cell(row=current_row + 2, column=3).value = "Stop Location"
        worksheet.cell(row=current_row + 3, column=3).value = "Onboard"

        # Display the time headers
        col = 5
        for datetime in sorted(self.times):
            worksheet.cell(row=current_row+1, column=col).value = \
                datetime.date()
            worksheet.cell(row=current_row+1, column=col+1).value = \
                datetime.time()
            worksheet.cell(row=current_row+2, column=col).value = "On"
            worksheet.cell(row=current_row+2, column=col+1).value = "Off"
            worksheet.cell(row=current_row+2, column=col+2).value = "OB"
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 11

            # Display onboard if it exists
            if datetime in self.onboard:
                worksheet.cell(row=current_row+3, column=col+2).value = \
                self.onboard[datetime]
            else:
                worksheet.cell(row=current_row+3, column=col+2).value = 0

            col += 3

        # Display the stops with info
        current_row += 4
        for stop_no in self.stops:
            self.stops[stop_no].buildDetailReport(worksheet, current_row)
            current_row += 1

        return current_row + 3
            

class Stop:
    """
    A class that represents a single stop on a bus route
    """
    def __init__(self, route, stop_no, street, cross_street) -> None:
        """
        Initialize with metadata
        
        @param route The route number this stop sits on
        @param stop_no The stop number of this stop
        @param street The main street of this stop
        @param cross_street The cross street of this stop
        """
        self.route = route
        self.stop_no = stop_no
        self.street = street
        self.cross_street = cross_street
        self.descriptor = "Descriptor Unset"
        self.direction = Direction.UN

        # Data is of the following format:
        # [
        #   run,
        #   arrival_time,
        #   schedule_time,
        #   offs,
        #   ons,
        #   load
        # ]
        self.data = {}

    def __str__(self) -> str:
        output = str(self.route) + ": " + str(self.stop_no) + " [" + \
            str(self.street) + "/" + str(self.cross_street) + "]\n"
        for datetime in sorted(self.data.keys()):
            output += str(datetime) + " " + str(self.data[datetime][0]) + " " + str(self.data[datetime][1]) + " " + str(self.data[datetime][2]) + " " + str(self.data[datetime][3]) + " " + str(self.data[datetime][4]) + "\n"
        output += "\n"
        return output

    def __repr__(self) -> str:
        return self.__str__()

    def addData(self, datetime, run, arrival_time, schedule_time, offs, ons)\
        -> bool:
        """
        Add data from a certain run to this stop
        
        @param datetime The datetime when this route begain
        @param run The run value of this stop
        @param arrival_time Optional: The arrival time for this stop
        @param schedule_time Optional: The scheduled arrival time for this stop
        @param offs Optional: The number of passengers departing the bus
        @param ons Optional: The number of passengers boarding the bus
        @returns Boolean of whether the data was successfully added
        """
        # Clean input data
        if not (isinstance(offs, int) and offs >= 0):
            offs = 0
        if not (isinstance(ons, int) and ons >= 0):
            ons = 0

        self.data[datetime] = [run, arrival_time, schedule_time, offs, ons, 0]
        return True

    def setLoad(self, datetime, load) -> None:
        """
        Sets the load for a given datetime.

        @param datetime The datetime when this route began
        @param load The passenger load
        """
        # Add data if doesn't exist
        if datetime not in self.data:
            self.data[datetime] = [None, None, None, 0, 0, load]
        else:
            self.data[datetime][5] = load

    def setRouteData(self, description, direction: Direction) -> None:
        """
        Sets the metadata for this route
        
        @param description A text description of the route (University, Uptown)
        @param direction The direction of the route as a Direction object
        """
        self.descriptor = description
        self.direction = direction

    def getDescriptorAndDirection(self) -> str:
        """
        @returns A string representation of this routes descriptor and direction
        """
        return self.descriptor + " " + str(self.direction.value)

    def getRun(self, datetime) -> str:
        # If datetime does not exist, return None
        if datetime not in self.data:
            return None

        return str(self.data[datetime][0])

    def getOffsAndOns(self, datetime) -> tuple:
        """
        @returns the offs and ons for a specific datetime
        """
        # If datetime does not exist, return zeros
        if datetime not in self.data:
            return 0, 0

        return self.data[datetime][3], self.data[datetime][4]

    def getOffsOnsAndLoad(self, datetime) -> tuple:
        """
        @returns the offs ons and load for a specific datetime
        """
        # If datetime does not exist, return zeros
        if datetime not in self.data:
            return 0, 0, 0

        return self.data[datetime][3], self.data[datetime][4],\
            self.data[datetime][5]

    def getTotalOffsAndOns(self) -> tuple:
        """
        @returns the total offs and ons for all datetimes at this stop
        """
        total_offs = 0
        total_ons = 0
        for datetime in self.data:
            total_offs += self.data[datetime][3]
            total_ons += self.data[datetime][4]
        return total_offs, total_ons

    def buildRouteTotalsByStop(self, worksheet, current_row) -> int:
        """
        Builds a route total for this stop for all datetimes

        @param worksheet The worksheet to operate on
        @param current_row The row to start on in the worksheet
        @returns The next blank row on the workbook
        """
        ons = 0
        offs = 0
        load = 0
        
        for key in self.data:
            offs += self.data[key][3]
            ons += self.data[key][4]
            load += self.data[key][5]
            
        total = ons + offs

        # Write data to sheet
        worksheet.cell(row=current_row, column=1).value = self.route
        worksheet.cell(row=current_row, column=2).value = \
            self.getDescriptorAndDirection()
        worksheet.cell(row=current_row, column=3).value = self.stop_no
        worksheet.cell(row=current_row, column=4).value = self.street
        worksheet.cell(row=current_row, column=5).value = self.cross_street
        worksheet.cell(row=current_row, column=6).value = ons
        worksheet.cell(row=current_row, column=7).value = offs
        worksheet.cell(row=current_row, column=8).value = total
        worksheet.cell(row=current_row, column=9).value = load

        return current_row + 1

    def getMinutesLate(self, datetime) -> int:
        """
        Gets the minutes late a bus is for a given datetime

        @param datetime The datetime of the stop
        @returns How many minutes the bus was late. Early busses are negatives.
            Returns None if datetime does not exist.
        """
        # If datetime doesn't exist, return None
        if datetime not in self.data:
            return None
        
        # If datetime does not have an arrival time or schedule time, ret None
        data = self.data[datetime]
        if data[1] is None or data[2] is None:
            return None

        # Data presumed good, return difference in minutes      
        delta = data[1] - data[2]
        return round(delta.total_seconds() / 60)

    def buildDetailReport(self, worksheet, current_row) -> None:
        """
        Builds a single row of a detail report sheet.

        @param worksheet The worksheet to modify
        @param current_row The current row being operated on
        """
        # Display header
        worksheet.cell(row=current_row, column=2).value = self.stop_no
        worksheet.cell(row=current_row, column=3).value = self.street
        worksheet.cell(row=current_row, column=4).value = self.cross_street

        # Display data for each datetime
        col = 5
        for datetime in sorted(self.data.keys()):
            worksheet.cell(row=current_row, column=col).value = \
                self.data[datetime][4]
            worksheet.cell(row=current_row, column=col + 1).value = \
                self.data[datetime][3]
            worksheet.cell(row=current_row, column=col + 2).value = \
                self.data[datetime][5]
            col += 3
        

def generateSummary(log:Log, ride_checks_filepath, bus_stop_filepath, 
    output_filepath) -> int:
    """
    Creates an output workbook from the provided input workbooks. See the README
    for more information on how this function operates.

    @param ride_checks_filepath The filepath for the ridechecks workbook
    @param bus_stop_filepath The filepath for the bus stop workbook
    @param output_filepath The filepath for the output workbook

    @returns An integer representing the status of the output workbook:
        0 - OK, success or minor errors
        1 - Major error. Check the workbook log for details.
        2 - Output workbook could not be created
    """

    # Create output workbook / sheet. Set date format to ISO 8601
    wb = openpyxl.Workbook()
    wb.iso_dates = True

    log.logGeneral("Output document created")

    # Try to open the ride checks file, if can't return major error
    log.logGeneral("Opening ride checks workbook")
    try:
        ride_checks_wb = openpyxl.load_workbook(filename=ride_checks_filepath, data_only=True)
    except Exception:
        log.logError("Could not open the ride checks workbook '" +\
            ride_checks_filepath + "'")
        
        # Try to save the output file
        try:
            wb.save(output_filepath)
        except Exception:
            return 2
        return 1
    ride_checks = ride_checks_wb.active

    # Try to open the bus stop file, if can't return major error
    log.logGeneral("Opening bus stop workbook")
    try:
        bus_stop_wb = openpyxl.load_workbook(filename=bus_stop_filepath, data_only=True)
    except Exception:
        log.logError("Could not open the bus stop workbook '" +\
            bus_stop_filepath + "'")
        
        # Try to save the output file
        try:
            wb.save(output_filepath)
        except Exception:
            return 2
        return 1
    bus_stop = bus_stop_wb.worksheets[0]

    route_manager = RouteManager(log)

    # parse the bus stop file to create route names and times
    log.logGeneral("Parsing bus stop file")
    current_route = None
    current_direction = None
    current_route_name = None

    for current_row in range(1,bus_stop.max_row + 1):
        # Check for new route header
        if bus_stop.cell(row=current_row, column=8).value == "ROUTE":
            # Get info of the new route
            current_route = bus_stop.cell(row=current_row, column=10).value
            current_route_name = \
                bus_stop.cell(row=current_row, column=4).value
            current_direction = stringToDirection(
                bus_stop.cell(row=current_row + 1, column=10).value)

            # Set the route data
            route_manager.setRouteData(current_route, current_route_name, current_direction)

        # If current route is None, skip this row
        if current_route is None:
            continue

        # If stop number is numberical, add this row
        stop_no = bus_stop.cell(row=current_row, column=5).value
        if isinstance(stop_no, int):
            street = bus_stop.cell(row=current_row, column=3).value
            cross_street = bus_stop.cell(row=current_row, column=4).value
            route_manager.addStop(current_route, stop_no, street, cross_street)

    # start parsing the ride checks file
    log.logGeneral("Parsing ride checks file")
    current_row = 2
    prev_seq = 0
    total_ons = 0
    total_offs = 0
    while(ride_checks.cell(row=current_row, column=1).value is not None):
        # get data
        sequence = ride_checks.cell(row=current_row, column=1).value
        date = ride_checks.cell(row=current_row, column=2).value
        route = ride_checks.cell(row=current_row, column=3).value
        direction = ride_checks.cell(row=current_row, column=4).value
        run = str(ride_checks.cell(row=current_row, column=5).value)
        start_time = ride_checks.cell(row=current_row, column=6).value
        onboard = ride_checks.cell(row=current_row, column=7).value
        stop_number = ride_checks.cell(row=current_row, column=8).value
        arrival_time = ride_checks.cell(row=current_row, column=9).value
        schedule_time = ride_checks.cell(row=current_row, column=10).value
        offs = ride_checks.cell(row=current_row, column=11).value
        ons = ride_checks.cell(row=current_row, column=12).value
        #loads = ride_checks.cell(row=current_row, column=13).value
        #time_check = ride_checks.cell(row=current_row, column=14).value

        # Check that the sequence number is in order, alert if not
        if sequence - 1 != prev_seq:
            log.logWarning("Out-of-order sequence number: Row " +
            str(current_row))
        prev_seq = sequence

        # check that all required data is the proper type
        if not isinstance(sequence, int):
            log.logError("Row " + str(current_row) + ": Sequence '" + \
                str(sequence) + "' is not an integer. Skipping row.")
            current_row += 1
            continue
        if not isinstance(date, datetime.date):
            log.logError("Row " + str(current_row) + ": Date '" + \
                str(date) + "' is not an excel-formatted date. " + \
                "Skipping row.")
            current_row += 1
            continue
        if not isinstance(route, int):
            log.logError("Row " + str(current_row) + ": Route '" + \
                str(route) + "' is not an integer. Skipping row.")
            current_row += 1
            continue
        if stringToDirection(direction) == Direction.UN:
            log.logError("Row " + str(current_row) + ": Direction '" + \
                str(direction) + "' is not a valid input. Skipping row.")
            current_row += 1
            continue 
        if (not isinstance(start_time, datetime.time)):
            log.logError("Row " + str(current_row) + ": Start time '" + \
                str(start_time) + "' is not an excel-formatted time. " + \
                "Skipping row.")
            current_row += 1
            continue

        # correct blank strings in optional data
        if arrival_time == "":
            arrival_time = None
        if schedule_time == "":
            schedule_time = None
        if ons == "":
            ons = None
        if offs == "":
            offs = None

        # check that all optional data is the correct format if filled in
        if (onboard is not None) and (not isinstance(onboard, int)):
            log.logError("Row " + str(current_row) + ": Onboard '" + \
                str(onboard) + "' is not an integer. Skipping row.")
            current_row += 1
            continue
        if (arrival_time is not None) and (not isinstance(arrival_time, \
            datetime.time)):
            log.logError("Row " + str(current_row) + ": Arrival time '" + \
                str(arrival_time) + "' is not an excel-formatted time."\
                + "Skipping row.")
            current_row += 1
            continue
        if (schedule_time is not None) and (not isinstance(schedule_time, \
            datetime.time)):
            log.logError("Row " + str(current_row) + ": Scheduled time '" + \
                str(schedule_time) + "' is not an excel-formatted time."\
                + "Skipping row.")
            current_row += 1
            continue
        if (ons is not None) and (not isinstance(ons, int)):
            log.logError("Row " + str(current_row) + ": Ons value '" + \
            str(ons) + "' is not an integer. Skipping row.")
            current_row += 1
            continue   
        if (offs is not None) and (not isinstance(offs, int)):
            log.logError("Row " + str(current_row) + ": Offs value '" + \
            str(offs) + "' is not an integer. Skipping row.")
            current_row += 1
            continue           

        # add data to the route manager object
        # order of placement is route string -> date and time -> stop_number
        start_datetime = datetime.datetime.combine(date, start_time)

        arrival_datetime = None
        if arrival_time is not None:
            arrival_datetime = datetime.datetime.combine(date, arrival_time)

        schedule_datetime = None
        if schedule_time is not None:
            schedule_datetime = datetime.datetime.combine(date, schedule_time)

        add_data_result = route_manager.addData(route, stop_number, \
            start_datetime, run, arrival_datetime, schedule_datetime, offs, \
            ons, onboard)

        if not add_data_result:
            log.logError("Row " + str(current_row) + ": Add data failure.")

        # increment total ons and offs
        if ons is not None:
            total_ons += ons
        if offs is not None:
            total_offs += offs

        # increment current row
        current_row += 1

    # Check for total ons and offs being equal
    if total_ons != total_offs:
        log.logWarning(f"Total ons and offs are not equal ({total_ons} ons, {total_offs} offs). Check for bad data")
    
    # Generate load data
    log.logGeneral("Building load data")
    route_manager.buildLoad()

    # DEBUG - Print all routes & stops & data
    # print(str(route_manager))

    # Remove existing sheets
    for s in wb.sheetnames:
        wb.remove(wb[s])

    # Generate route totals sheet
    log.logGeneral("Generating route totals")
    routeTotalsSheet = wb.create_sheet("Rte Totals")
    route_manager.buildRouteTotals(routeTotalsSheet)

    # Generate max load sheet
    log.logGeneral("Generating max load sheet")
    maxLoadSheet = wb.create_sheet("Max Load")
    route_manager.buildMaxLoads(maxLoadSheet)

    # Generate totals by stop sheet
    log.logGeneral("Generating route totals per stop")
    maxLoadSheet = wb.create_sheet("Ons Offs Tot & Ld")
    route_manager.buildRouteTotalsByStop(maxLoadSheet)

    # Generate the on-time detail
    log.logGeneral("Generating on-time detail")
    onTimeSheet = wb.create_sheet("On Time Detail")
    route_manager.buildOnTimeDetail(onTimeSheet)

    # Generate the detail report
    log.logGeneral("Generating detail report")
    detailReportSheet = wb.create_sheet("Detail Report")
    route_manager.buildDetailReport(detailReportSheet)

    # Add notes sheet
    wb.create_sheet("Notes")

    # Generation complete
    log.logGeneral("Generation complete")

    # Try to save the output file
    try:
        wb.save(output_filepath)
    except Exception:
        return 2

    # Return successfully
    return 0
    
